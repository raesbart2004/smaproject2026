"""
Comparison design with secondary objectives.

Save this file as:
    comparison_design_secondary_objectives.py

Put it in:
    smaproject2026/python-code/

This file does NOT overwrite your previous comparison design code.
It imports the Simulation class from your existing comparison_design.py file.
"""

import os
import math
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from comparison_design import Simulation


# ============================================================
# SETTINGS
# ============================================================

W = 500
R = 12

slots_to_test = [12, 13]
rules_to_test = [2, 4]
strategies_to_test = [2, 3]

# All-pairwise OV:
# n = 12, df = 11, Bonferroni correction for 28 comparisons, family confidence 90%
T_VALUE_PAIRWISE_OV = 3.704

# Secondary objectives:
# n = 12, df = 11, 90% paired t-test
T_VALUE_SECONDARY = 1.796

# Remaining candidates after all-pairwise comparison.
secondary_comparisons = [
    ("X3", "X2"),
    ("X3", "X7"),
]


def get_input_path(strategy, slots):
    return f"../input-S{strategy}-{slots}.txt"


def avg(values):
    return sum(values) / len(values)


def sample_std(values):
    m = avg(values)
    return math.sqrt(sum((x - m) ** 2 for x in values) / (len(values) - 1))


def run_simulation_with_secondary_outputs(sim):
    """
    Runs the simulation and returns:
        OV, EAWT, ESWT, USWT, OT

    OV   = objective value
    EAWT = elective appointment waiting time
    ESWT = elective scan waiting time
    USWT = urgent scan waiting time
    OT   = overtime

    This keeps the same antithetic variables logic as your original runSimulations().
    """

    sim.setWeekSchedule()

    total_OV = 0.0
    total_EAWT = 0.0
    total_ESWT = 0.0
    total_USWT = 0.0
    total_OT = 0.0

    for _ in range(sim.R):
        start_seed = random.randint(0, 1_000_000)

        # NORMAL RUN
        random.seed(start_seed)
        sim.resetSystem()
        sim.generatePatients(is_antithetic=False)
        sim.runOneSimulation()

        norm_EAWT = sim.avgElectiveAppWT
        norm_ESWT = sim.avgElectiveScanWT
        norm_USWT = sim.avgUrgentScanWt
        norm_OT = sim.avgOT
        norm_OV = (norm_EAWT * sim.weightEl) + (norm_USWT * sim.weightUr)

        # ANTITHETIC RUN
        random.seed(start_seed)
        sim.resetSystem()
        sim.generatePatients(is_antithetic=True)
        sim.runOneSimulation()

        anti_EAWT = sim.avgElectiveAppWT
        anti_ESWT = sim.avgElectiveScanWT
        anti_USWT = sim.avgUrgentScanWt
        anti_OT = sim.avgOT
        anti_OV = (anti_EAWT * sim.weightEl) + (anti_USWT * sim.weightUr)

        # COMBINE NORMAL + ANTITHETIC
        total_OV += (norm_OV + anti_OV) / 2
        total_EAWT += (norm_EAWT + anti_EAWT) / 2
        total_ESWT += (norm_ESWT + anti_ESWT) / 2
        total_USWT += (norm_USWT + anti_USWT) / 2
        total_OT += (norm_OT + anti_OT) / 2

    return (
        total_OV / sim.R,
        total_EAWT / sim.R,
        total_ESWT / sim.R,
        total_USWT / sim.R,
        total_OT / sim.R,
    )


def paired_ci(values_i, values_j, t_value):
    """
    Calculates paired confidence interval for Xi - Xj.
    """

    n = len(values_i)
    differences = [values_i[r] - values_j[r] for r in range(n)]

    mean_diff = avg(differences)
    sd_diff = sample_std(differences)
    half_width = t_value * sd_diff / math.sqrt(n)

    ci_low = mean_diff - half_width
    ci_high = mean_diff + half_width

    if ci_high < 0:
        conclusion = "Xi significantly better than Xj"
    elif ci_low > 0:
        conclusion = "Xj significantly better than Xi"
    else:
        conclusion = "No significant difference"

    return mean_diff, sd_diff, half_width, ci_low, ci_high, conclusion


def add_pairwise_sheet(wb, sheet_name, design_points, results_dict, t_value):
    ws = wb.create_sheet(sheet_name)

    ws.append([
        "Pair",
        "Mean difference Xi-Xj",
        "Std Dev difference",
        "t-value",
        "Half-width",
        "CI-",
        "CI+",
        "Conclusion",
    ])

    for i in range(len(design_points)):
        for j in range(i + 1, len(design_points)):
            Xi = design_points[i]["DP"]
            Xj = design_points[j]["DP"]

            mean_diff, sd_diff, half_width, ci_low, ci_high, conclusion = paired_ci(
                results_dict[Xi],
                results_dict[Xj],
                t_value,
            )

            if conclusion == "Xi significantly better than Xj":
                conclusion_text = f"{Xi} significantly better than {Xj}"
            elif conclusion == "Xj significantly better than Xi":
                conclusion_text = f"{Xj} significantly better than {Xi}"
            else:
                conclusion_text = "No significant difference"

            ws.append([
                f"{Xi} - {Xj}",
                mean_diff,
                sd_diff,
                t_value,
                half_width,
                ci_low,
                ci_high,
                conclusion_text,
            ])


def add_table_f_sheet(wb, design_points, results_ov):
    ws = wb.create_sheet("Table_F_OV")

    headers = ["i/j"] + [str(i) for i in range(2, 9)]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, label in enumerate([str(i) for i in range(1, 8)], start=2):
        cell = ws.cell(row=row_idx, column=1, value=label)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    green_font = Font(color="008000")
    thin = Side(style="thin", color="000000")

    for i in range(len(design_points) - 1):
        for j in range(i + 1, len(design_points)):
            Xi = design_points[i]["DP"]
            Xj = design_points[j]["DP"]

            mean_diff, _, half_width, ci_low, ci_high, _ = paired_ci(
                results_ov[Xi],
                results_ov[Xj],
                T_VALUE_PAIRWISE_OV,
            )

            significant = not (ci_low <= 0 <= ci_high)

            text = f"{mean_diff:.5f}±{half_width:.5f}"
            if significant:
                text += "*"

            excel_row = i + 2
            excel_col = j + 1

            cell = ws.cell(row=excel_row, column=excel_col, value=text)
            cell.alignment = Alignment(horizontal="center")

            if significant:
                cell.font = green_font

    for row in ws.iter_rows(min_row=1, max_row=8, min_col=1, max_col=8):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 22


def add_secondary_objectives_sheet(wb, comparisons, results_eswt, results_ot):
    ws = wb.create_sheet("Secondary_Objectives")

    ws.append([
        "Objective",
        "Comparison",
        "t-value",
        "Mean difference",
        "Half length",
        "CI-",
        "CI+",
        "Conclusion",
    ])

    for objective_name, data_dict in [
        ("ESWT", results_eswt),
        ("OT", results_ot),
    ]:
        for Xi, Xj in comparisons:
            mean_diff, _, half_width, ci_low, ci_high, conclusion = paired_ci(
                data_dict[Xi],
                data_dict[Xj],
                T_VALUE_SECONDARY,
            )

            if conclusion == "Xi significantly better than Xj":
                conclusion_text = f"{Xi} significantly better than {Xj}"
            elif conclusion == "Xj significantly better than Xi":
                conclusion_text = f"{Xj} significantly better than {Xi}"
            else:
                conclusion_text = "No significant difference"

            ws.append([
                objective_name,
                f"{Xi} - {Xj}",
                T_VALUE_SECONDARY,
                mean_diff,
                half_width,
                ci_low,
                ci_high,
                conclusion_text,
            ])


def main():
    if not os.path.exists("Resultaten"):
        os.makedirs("Resultaten")

    output_filename = os.path.join(
        "Resultaten",
        "Comparison_Design_Secondary_Objectives.xlsx"
    )

    design_points = []
    dp = 1

    for slots in slots_to_test:
        for rule in rules_to_test:
            for strategy in strategies_to_test:
                design_points.append({
                    "DP": f"X{dp}",
                    "Slots": slots,
                    "Rule": rule,
                    "Strategy": strategy,
                })
                dp += 1

    missing_files = []

    for d in design_points:
        input_path = get_input_path(d["Strategy"], d["Slots"])
        if not os.path.exists(input_path):
            missing_files.append(input_path)

    if missing_files:
        print("\nERROR: Missing input files:")
        for file in missing_files:
            print(" -", file)
        raise FileNotFoundError("Some required input files are missing.")

    wb = openpyxl.Workbook()

    ws_ov = wb.active
    ws_ov.title = "Raw_OV"
    ws_ov.append(["Replication"] + [d["DP"] for d in design_points])

    ws_eawt = wb.create_sheet("Raw_EAWT")
    ws_eawt.append(["Replication"] + [d["DP"] for d in design_points])

    ws_eswt = wb.create_sheet("Raw_ESWT")
    ws_eswt.append(["Replication"] + [d["DP"] for d in design_points])

    ws_uswt = wb.create_sheet("Raw_USWT")
    ws_uswt.append(["Replication"] + [d["DP"] for d in design_points])

    ws_ot = wb.create_sheet("Raw_OT")
    ws_ot.append(["Replication"] + [d["DP"] for d in design_points])

    results_ov = {d["DP"]: [] for d in design_points}
    results_eawt = {d["DP"]: [] for d in design_points}
    results_eswt = {d["DP"]: [] for d in design_points}
    results_uswt = {d["DP"]: [] for d in design_points}
    results_ot = {d["DP"]: [] for d in design_points}

    print("\n" + "=" * 80)
    print("RUNNING COMPARISON DESIGN WITH SECONDARY OBJECTIVES")
    print(f"R = {R}, W = {W}")
    print("=" * 80)

    for r in range(1, R + 1):
        row_ov = [r]
        row_eawt = [r]
        row_eswt = [r]
        row_uswt = [r]
        row_ot = [r]

        current_seed = 1000 + r

        for d in design_points:
            input_path = get_input_path(d["Strategy"], d["Slots"])
            rule = d["Rule"]

            random.seed(current_seed)

            sim = Simulation(input_path, W, 1, rule)

            ov, eawt, eswt, uswt, ot = run_simulation_with_secondary_outputs(sim)

            dp_name = d["DP"]

            results_ov[dp_name].append(ov)
            results_eawt[dp_name].append(eawt)
            results_eswt[dp_name].append(eswt)
            results_uswt[dp_name].append(uswt)
            results_ot[dp_name].append(ot)

            row_ov.append(ov)
            row_eawt.append(eawt)
            row_eswt.append(eswt)
            row_uswt.append(uswt)
            row_ot.append(ot)

        ws_ov.append(row_ov)
        ws_eawt.append(row_eawt)
        ws_eswt.append(row_eswt)
        ws_uswt.append(row_uswt)
        ws_ot.append(row_ot)

        print(f"Replication {r}/{R} finished")

    ws_design = wb.create_sheet("Design_Points")
    ws_design.append(["Design point", "#Slots", "Rule", "Strategy"])

    for d in design_points:
        ws_design.append([d["DP"], d["Slots"], d["Rule"], d["Strategy"]])

    ws_summary = wb.create_sheet("Summary")
    ws_summary.append([
        "Design point",
        "#Slots",
        "Rule",
        "Strategy",
        "Mean OV",
        "Mean EAWT",
        "Mean ESWT",
        "Mean USWT",
        "Mean OT",
    ])

    ranking = []

    for d in design_points:
        dp_name = d["DP"]

        avg_ov = avg(results_ov[dp_name])
        avg_eawt = avg(results_eawt[dp_name])
        avg_eswt = avg(results_eswt[dp_name])
        avg_uswt = avg(results_uswt[dp_name])
        avg_ot = avg(results_ot[dp_name])

        ws_summary.append([
            dp_name,
            d["Slots"],
            d["Rule"],
            d["Strategy"],
            avg_ov,
            avg_eawt,
            avg_eswt,
            avg_uswt,
            avg_ot,
        ])

        ranking.append((avg_ov, d))

    ranking.sort(key=lambda x: x[0])

    ws_rank = wb.create_sheet("Ranking")
    ws_rank.append(["Rank", "Design point", "#Slots", "Rule", "Strategy", "Mean OV"])

    for rank, (mean_ov, d) in enumerate(ranking, start=1):
        ws_rank.append([rank, d["DP"], d["Slots"], d["Rule"], d["Strategy"], mean_ov])

    add_pairwise_sheet(wb, "Pairwise_OV", design_points, results_ov, T_VALUE_PAIRWISE_OV)
    add_pairwise_sheet(wb, "Pairwise_ESWT", design_points, results_eswt, T_VALUE_SECONDARY)
    add_pairwise_sheet(wb, "Pairwise_OT", design_points, results_ot, T_VALUE_SECONDARY)

    add_table_f_sheet(wb, design_points, results_ov)

    add_secondary_objectives_sheet(
        wb,
        secondary_comparisons,
        results_eswt,
        results_ot,
    )

    wb.save(output_filename)

    print("\n" + "=" * 80)
    print("DONE")
    print(f"Saved in: {output_filename}")
    print("=" * 80)

    best_avg, best_design = ranking[0]
    print("\nBest design point based on mean OV:")
    print(
        f"{best_design['DP']} | "
        f"Slots={best_design['Slots']} | "
        f"Rule={best_design['Rule']} | "
        f"Strategy={best_design['Strategy']} | "
        f"Mean OV={best_avg:.6f}"
    )


if __name__ == "__main__":
    main()
